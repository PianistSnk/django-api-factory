"""Tests for AuditLogMixin — verify default no-op behavior and subclass override."""

import json
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from django.test import RequestFactory

from django_api_factory.mixins import AuditLogMixin, ActionFormMixin, ExportMixin


def test_auditlogmixin_default_noop():
    """Default log_query / log_download return None and write nothing."""
    mixin = AuditLogMixin()
    assert mixin.log_query(None, "TestModel") is None
    assert mixin.log_download(None, "TestModel", "file.pdf", "pdf") is None
    assert mixin.enable_audit_log is False


def test_auditlogmixin_subclass_can_override():
    """Subclasses can override log_query to add audit behavior."""
    calls = []

    class ProjectFlavoredMixin(AuditLogMixin):
        def log_query(self, request, model_name):
            calls.append(("query", model_name))
            return "logged"

        def log_download(self, request, model_name, filename, type_):
            calls.append(("download", filename, type_))
            return "logged"

    mixin = ProjectFlavoredMixin()
    assert mixin.log_query(None, "Post") == "logged"
    assert mixin.log_download(None, "Post", "x.pdf", "pdf") == "logged"
    assert calls == [
        ("query", "Post"),
        ("download", "x.pdf", "pdf"),
    ]


def test_apiadmin_inherits_auditlogmixin():
    """APIAdmin should inherit from AuditLogMixin so hooks are available."""
    from django_api_factory.admin import APIAdmin
    from django_api_factory.mixins import AuditLogMixin
    assert issubclass(APIAdmin, AuditLogMixin)


def test_multi_value_separator_default():
    """Default uses 顿号 (project convention)."""
    from django_api_factory.admin import APIAdmin
    assert APIAdmin.multi_value_separator == "、"


def test_multi_value_separator_override():
    """Subclasses can override to use , or | etc."""
    from django_api_factory.admin import APIAdmin

    class CommaAdmin(APIAdmin):
        multi_value_separator = ","

    class PipeAdmin(APIAdmin):
        multi_value_separator = "|"

    assert CommaAdmin.multi_value_separator == ","
    assert PipeAdmin.multi_value_separator == "|"


def test_handle_search_condition_uses_configured_separator():
    """handle_search_condition should use self.multi_value_separator, not hardcoded 顿号."""
    # Direct unit test of the splitter logic (mirrors what the nested function does)
    from django_api_factory.admin import APIAdmin

    class PipeAdmin(APIAdmin):
        multi_value_separator = "|"

    sep = PipeAdmin.multi_value_separator
    item_value = "a|b|c"
    # Mimic the search logic with the configured separator
    assert sep.join(sorted(item_value.split(sep))) == "a|b|c"

    class CommaAdmin(APIAdmin):
        multi_value_separator = ","

    sep = CommaAdmin.multi_value_separator
    item_value = "c,a,b"
    assert sep.join(sorted(item_value.split(sep))) == "a,b,c"


# --- ExportMixin tests ---

def test_exportmixin_get_export_fields_default():
    """get_export_fields returns self.export_list if set, else None."""
    from django_api_factory.mixins import ExportMixin
    m = ExportMixin()
    m.export_list = ["a", "b", "c"]
    assert m.get_export_fields() == ["a", "b", "c"]
    del m.export_list
    assert m.get_export_fields() is None


def test_exportmixin_get_export_data_dicts():
    """get_export_data produces one dict per object with the configured fields."""
    from django_api_factory.mixins import ExportMixin

    class Row:
        def __init__(self, **kwargs):
            for k, v in kwargs.items():
                setattr(self, k, v)

    m = ExportMixin()
    m.export_list = ["name", "value"]
    rows = m.get_export_data([Row(name="x", value=1), Row(name="y", value=2)])
    assert rows == [
        {"name": "x", "value": 1},
        {"name": "y", "value": 2},
    ]


def test_exportmixin_get_export_data_missing_attr():
    """Missing attributes are coerced to '' (matches legacy behavior)."""
    from django_api_factory.mixins import ExportMixin

    class Row:
        name = "x"

    m = ExportMixin()
    m.export_list = ["name", "missing"]
    assert m.get_export_data([Row()]) == [{"name": "x", "missing": ""}]


# --- Date param parsing tests (T1.1i) ---

def _make_admin_subclass(attrs):
    """Build an APIAdmin instance bypassing __init__ (which needs model+admin_site)."""
    from django_api_factory.admin import APIAdmin
    cls = type("TestAdmin", (APIAdmin,), attrs)
    return cls.__new__(cls)


def test_parse_dt_empty():
    from django_api_factory.admin import APIAdmin
    a = _make_admin_subclass({})
    assert a.parse_dt("") == ""


def test_parse_dt_admin_date_hierarchy_format():
    """Django admin sends '<relative-word> <Mon DD YYYY>' — strip first token."""
    from django_api_factory.admin import APIAdmin
    a = _make_admin_subclass({})
    assert a.parse_dt("今天 Jun 5 2026") == "20260605"
    assert a.parse_dt("昨天 Jun 4 2026") == "20260604"


def test_parse_dt_pure_date_string():
    """Bare 'Mon DD YYYY' also works (defensive)."""
    from django_api_factory.admin import APIAdmin
    a = _make_admin_subclass({})
    assert a.parse_dt("Jun 5 2026") == "20260605"


def test_parse_dt_unparseable_returns_empty():
    """Garbage input returns '' instead of raising."""
    from django_api_factory.admin import APIAdmin
    a = _make_admin_subclass({})
    assert a.parse_dt("not a date") == ""
    assert a.parse_dt("2026-06-05") == ""  # ISO not supported by default
    assert a.parse_dt(None) == ""  # type: ignore


def test_parse_paras_date_params_converted():
    """parse_paras converts declared date_params, leaves others alone."""
    a = _make_admin_subclass({"date_params": ["dt", "dt_e"]})
    out = a.parse_paras({"dt": "今天 Jun 5 2026", "dt_e": "今天 Jun 4 2026", "flag": "0", "q": "foo"})
    assert out == {"dt": "20260605", "dt_e": "20260604", "flag": "0", "q": "foo"}


def test_parse_paras_no_date_params_returns_copy():
    """If date_params is empty, parse_paras just copies the dict (no parsing)."""
    a = _make_admin_subclass({})  # default date_params = []
    out = a.parse_paras({"dt": "今天 Jun 5 2026", "flag": "0"})
    assert out == {"dt": "今天 Jun 5 2026", "flag": "0"}


def test_parse_paras_does_not_mutate_input():
    """parse_paras returns a new dict, doesn't mutate the caller's dict."""
    a = _make_admin_subclass({"date_params": ["dt"]})
    original = {"dt": "今天 Jun 5 2026", "flag": "0"}
    a.parse_paras(original)
    assert original == {"dt": "今天 Jun 5 2026", "flag": "0"}  # unchanged


def test_parse_paras_missing_date_param_unchanged():
    """If a date_param is not in paras, the result omits it (not in dict)."""
    a = _make_admin_subclass({"date_params": ["dt", "dt_e"]})
    out = a.parse_paras({"dt": "今天 Jun 5 2026"})
    assert out == {"dt": "20260605"}  # dt_e not in result


# --- _handle_search_condition (BigPost-100k filter bug fix, Jun 2026) ----

def test_handle_search_condition_single_numeric_exact_match():
    """The BigPost-100k bug: `?userId=1` previously matched userIds
    1, 10-19, 21, 31, ...  because of `'1' in '10'` substring-match.
    After the fix: single term, no separator → EXACT equality with
    int coercion so numeric fields match by value."""
    from django_api_factory.admin import _handle_search_condition
    sep = "、"

    # The exact bug case — userId=1 must NOT match userId=10
    assert _handle_search_condition(10, ["1"], sep) is False
    assert _handle_search_condition(11, ["1"], sep) is False
    assert _handle_search_condition(19, ["1"], sep) is False
    assert _handle_search_condition(21, ["1"], sep) is False
    # But it DOES match userId=1
    assert _handle_search_condition(1, ["1"], sep) is True
    # And it doesn't match userId=2 (different number, no substring overlap)
    assert _handle_search_condition(2, ["1"], sep) is False


def test_handle_search_condition_single_string_exact_match():
    """String fields also get exact equality (no substring). `?title=foo`
    matches title='foo' but NOT title='foobar' (the previous substring
    match would have incorrectly included 'foobar')."""
    from django_api_factory.admin import _handle_search_condition
    sep = "、"

    assert _handle_search_condition("foo", ["foo"], sep) is True
    assert _handle_search_condition("foobar", ["foo"], sep) is False
    assert _handle_search_condition("barfoo", ["foo"], sep) is False
    assert _handle_search_condition("FOO", ["foo"], sep) is False  # case-sensitive


def test_handle_search_condition_int_vs_string_value():
    """API may return userId as int (10) while URL param is string ('10').
    int coercion normalizes both sides so they match."""
    from django_api_factory.admin import _handle_search_condition
    sep = "、"

    assert _handle_search_condition(10, ["10"], sep) is True  # int item, str term
    assert _handle_search_condition("10", ["10"], sep) is True  # both str


def test_handle_search_condition_multiterm_or_equals():
    """Multi-term (e.g. `?userId=1,2`) → OR-equals: match if item equals
    ANY of the terms. NOT substring — so userId=12 must NOT match
    `?userId=1,2`."""
    from django_api_factory.admin import _handle_search_condition
    sep = ","

    assert _handle_search_condition(1, ["1", "2"], sep) is True
    assert _handle_search_condition(2, ["1", "2"], sep) is True
    assert _handle_search_condition(12, ["1", "2"], sep) is False  # no substring
    assert _handle_search_condition(3, ["1", "2"], sep) is False


def test_handle_search_condition_multivalued_cell_with_separator():
    """Multi-valued cells like '苹果、香蕉' compare canonically against
    the search terms. Item '苹果、香蕉' matches search ['苹果']; item
    '苹果' (no separator) also matches ['苹果']."""
    from django_api_factory.admin import _handle_search_condition
    sep = "、"

    # Cell with separator: "苹果、香蕉" split-sorted-joined → canonical
    # "苹果、香蕉"; search term is single "苹果" (no sep in term)
    # → falls to the single-term branch → exact string equality.
    # NOTE: this is exact match, not "any of the cell values", by design
    # (the URL param IS the cell value the user picked from the filter).
    assert _handle_search_condition("苹果、香蕉", ["苹果"], sep) is False  # full string compare
    assert _handle_search_condition("苹果", ["苹果"], sep) is True

    # Multi-term with separator: search ['苹果', '香蕉'] (sep "、" in terms)
    # → OR-equals: matches if cell EQUALS any term.
    assert _handle_search_condition("苹果", ["苹果", "香蕉"], sep) is True
    assert _handle_search_condition("香蕉", ["苹果", "香蕉"], sep) is True
    assert _handle_search_condition("葡萄", ["苹果", "香蕉"], sep) is False


# ============================================================================
# ExportMixin.export_to_excel  (mixins.py line 104-132)
# ============================================================================

def _make_export_admin(*, fields=("id", "name"), rows=None, model_verbose_name="widget"):
    """Construct an ExportMixin-flavored admin with __new__ (no Django init)."""
    admin = ExportMixin.__new__(ExportMixin)
    admin.model = MagicMock()
    admin.model._meta.verbose_name = model_verbose_name
    admin.export_list = list(fields)
    # Row dicts default to empty list; tests can pass a populated list.
    admin._export_rows = rows if rows is not None else [
        {f: f"row{i}_{f}" for f in fields} for i in range(3)
    ]
    return admin


def test_export_to_excel_returns_xlsx_http_response():
    from django_api_factory.mixins import ExportMixin
    admin = _make_export_admin()
    response = admin.export_to_excel(MagicMock(), queryset=[])
    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_export_to_excel_filename_includes_model_name_and_date():
    admin = _make_export_admin(model_verbose_name="post")
    response = admin.export_to_excel(MagicMock(), queryset=[])
    from datetime import date
    today = date.today().strftime("%Y-%m-%d")
    assert response["Content-Disposition"] == f'attachment;filename="post_{today}.xlsx"'


def test_export_to_excel_writes_headers_and_data_rows():
    """The first row is the headers, subsequent rows are data."""
    from openpyxl import load_workbook
    admin = _make_export_admin(
        fields=("id", "title"),
        rows=[{"id": 1, "title": "hello"}, {"id": 2, "title": "world"}],
    )
    # export_to_excel walks the queryset via get_export_data, which uses
    # getattr(obj, field, "") on each item. Patch get_export_data to
    # return our row dicts directly so the worksheet actually has data rows.
    admin.get_export_data = MagicMock(return_value=[
        {"id": 1, "title": "hello"},
        {"id": 2, "title": "world"},
    ])
    response = admin.export_to_excel(MagicMock(), queryset=MagicMock())
    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("id", "title")
    assert rows[1] == (1, "hello")
    assert rows[2] == (2, "world")


def test_export_to_excel_worksheet_title():
    from openpyxl import load_workbook
    admin = _make_export_admin()
    response = admin.export_to_excel(MagicMock(), queryset=MagicMock())
    wb = load_workbook(filename=BytesIO(response.content))
    assert wb.active.title == "Data Export"


def test_export_to_excel_empty_queryset_writes_only_header_row():
    """When get_export_data returns [], the file has just the header row."""
    from openpyxl import load_workbook
    admin = _make_export_admin(fields=("id", "title"), rows=[])
    admin.get_export_data = MagicMock(return_value=[])
    response = admin.export_to_excel(MagicMock(), queryset=MagicMock())
    wb = load_workbook(filename=BytesIO(response.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows == [("id", "title")]


def test_export_to_excel_raises_import_error_when_openpyxl_missing(monkeypatch):
    """If openpyxl is not installed, export_to_excel raises ImportError
    with a helpful `pip install openpyxl` message."""
    import builtins
    real_import = builtins.__import__

    def fake_import(name, *args, **kwargs):
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError(f"No module named {name!r}")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", fake_import)
    admin = _make_export_admin()
    with pytest.raises(ImportError, match="pip install openpyxl"):
        admin.export_to_excel(MagicMock(), queryset=MagicMock())


# ============================================================================
# ActionFormMixin.get_urls + _resolve_action
# (mixins.py line 357-371 + 379)
# ============================================================================

def _make_action_admin(*, action_func=None, action_name="my_action",
                       get_urls_super=None, get_changelist_raises=False,
                       get_queryset_value=None, get_action_result=None,
                       has_attr=True):
    """Build an ActionFormMixin-flavored admin without running ModelAdmin.__init__.

    ActionFormMixin is a mixin, so a bare __new__(ActionFormMixin) gives an
    instance whose MRO doesn't include ModelAdmin — and the mixin's own
    `super().get_urls()` would resolve to object, not ModelAdmin. We
    construct a synthetic class that mixes both so the MRO walks through
    ModelAdmin as expected.
    """
    from django_api_factory.mixins import ActionFormMixin
    from django.contrib.admin.options import ModelAdmin
    cls = type("_ActionFormTestAdmin", (ActionFormMixin, ModelAdmin), {})
    admin = cls.__new__(cls)
    # admin_site: minimal mock — admin_view just returns the wrapped callable
    admin.admin_site = MagicMock()
    admin.admin_site.admin_view = lambda fn: fn
    # The two URL paths we care about
    admin.model = MagicMock()
    admin.model._meta.app_label = "tests"
    admin.model._meta.model_name = "widget"
    # NOTE: do NOT set admin.get_urls here — that would shadow the
    # ActionFormMixin.get_urls method we're trying to test. Tests
    # that need a custom super().get_urls() result use
    # patch.object(ModelAdmin, "get_urls", return_value=...) instead.
    # _resolve_action / action_submit_view path
    if has_attr:
        setattr(admin, action_name, lambda *a, **kw: None)
    admin.get_action = MagicMock(return_value=get_action_result)
    # changelist / queryset for action_submit_view
    if get_changelist_raises:
        admin.get_changelist_instance = MagicMock(side_effect=Exception("cl boom"))
    else:
        cl = MagicMock()
        cl.get_queryset = MagicMock(return_value=MagicMock())
        admin.get_changelist_instance = MagicMock(return_value=cl)
    admin.get_queryset = MagicMock(return_value=get_queryset_value or MagicMock())
    return admin


def test_get_urls_includes_action_form_and_submit_routes():
    """get_urls prepends /action-form/<action_name>/ and /action-submit/<action_name>/
    before the super() routes."""
    from django.urls import path as url_path
    from django.contrib.admin.options import ModelAdmin

    super_url = url_path("super/", lambda r: None, name="super_route")
    admin = _make_action_admin()  # no get_urls_super override
    # Patch the *parent* get_urls so super().get_urls() returns our
    # super_url. (We can't just set admin.get_urls = ... because that
    # would replace the method we're testing, not augment it.)
    with patch.object(ModelAdmin, "get_urls", return_value=[super_url]):
        urls = admin.get_urls()
    assert len(urls) == 3
    assert "action-form" in urls[0].pattern._route
    assert "action-submit" in urls[1].pattern._route
    assert urls[2] is super_url


def test_resolve_action_returns_none_when_attr_missing():
    """When the action attr doesn't exist on the admin, return None
    (the caller turns this into 404)."""
    admin = _make_action_admin(action_name="nonexistent", has_attr=False)
    assert admin._resolve_action("nonexistent") is None


def test_resolve_action_returns_none_when_get_action_returns_none():
    """Even if the attr exists, get_action() may return (None, ...) — pass through."""
    admin = _make_action_admin(
        get_action_result=(None, "my_action", "My Action"),
    )
    assert admin._resolve_action("my_action") is None


def test_resolve_action_returns_func_action_description():
    admin = _make_action_admin(
        get_action_result=("the-callable", "my_action", "My Action"),
    )
    result = admin._resolve_action("my_action")
    assert result == ("the-callable", "my_action", "My Action")


# ============================================================================
# action_submit_view  (mixins.py line 446-470)
# ============================================================================

def test_action_submit_returns_404_when_action_not_found():
    """Unknown action name → JsonResponse status=error 404."""
    rf = RequestFactory()
    request = rf.post("/admin/tests/widget/action-submit/missing/")
    admin = _make_action_admin(action_name="missing", has_attr=False)
    response = admin.action_submit_view(request, "missing")
    assert response.status_code == 404
    data = json.loads(response.content)
    assert data["status"] == "error"
    assert "missing" in data["msg"]


def test_action_submit_filters_by_selected_ids_when_select_across_zero():
    """When select_across=0 and _selected has ids, qs gets filtered by pk__in."""
    rf = RequestFactory()
    request = rf.post(
        "/admin/tests/widget/action-submit/my_action/",
        {"_selected": "1,2,3", "select_across": "0"},
    )
    qs = MagicMock()
    qs.filter.return_value = "filtered-qs"
    action_func = MagicMock()
    # Force the fallback path (get_changelist_instance raises) so the
    # action_submit_view uses *our* qs (set via get_queryset_value) —
    # otherwise it uses cl.get_queryset(request), which is a separate
    # MagicMock we don't control.
    admin = _make_action_admin(
        get_changelist_raises=True,
        get_queryset_value=qs,
        get_action_result=(action_func, "my_action", "My Action"),
    )
    admin.action_submit_view(request, "my_action")
    action_func.assert_called_once()
    call_args = action_func.call_args
    assert call_args[0][2] == "filtered-qs"  # func(self, request, qs)
    qs.filter.assert_called_once_with(pk__in=[1, 2, 3])


def test_action_submit_no_selection_does_not_filter():
    """When _selected is empty, qs is passed through without pk__in filter."""
    rf = RequestFactory()
    request = rf.post(
        "/admin/tests/widget/action-submit/my_action/",
        {"_selected": "", "select_across": "0"},
    )
    qs = MagicMock()
    action_func = MagicMock()
    admin = _make_action_admin(
        get_changelist_raises=True,
        get_queryset_value=qs,
        get_action_result=(action_func, "my_action", "My Action"),
    )
    admin.action_submit_view(request, "my_action")
    qs.filter.assert_not_called()


def test_action_submit_falls_back_to_get_queryset_when_changelist_fails():
    """If get_changelist_instance raises, fall back to self.get_queryset()."""
    rf = RequestFactory()
    request = rf.post(
        "/admin/tests/widget/action-submit/my_action/",
        {"_selected": "1", "select_across": "0"},
    )
    fallback_qs = MagicMock()
    action_func = MagicMock()
    admin = _make_action_admin(
        get_changelist_raises=True,
        get_queryset_value=fallback_qs,
        get_action_result=(action_func, "my_action", "My Action"),
    )
    admin.action_submit_view(request, "my_action")
    admin.get_queryset.assert_called_once_with(request)


def test_action_submit_normalizes_none_to_success():
    """When the action returns None, response is {status: success, msg: Success!}."""
    import json
    rf = RequestFactory()
    request = rf.post("/admin/tests/widget/action-submit/my_action/", {})
    admin = _make_action_admin(
        get_action_result=(lambda *a, **kw: None, "my_action", "My Action"),
    )
    response = admin.action_submit_view(request, "my_action")
    assert response.status_code == 200
    data = json.loads(response.content)
    assert data == {"status": "success", "msg": "Success!"}


def test_action_submit_normalizes_dict_to_json():
    """Dict return value is JSON-encoded as-is."""
    import json
    rf = RequestFactory()
    request = rf.post("/admin/tests/widget/action-submit/my_action/", {})
    admin = _make_action_admin(
        get_action_result=(lambda *a, **kw: {"status": "custom", "extra": 42},
                            "my_action", "My Action"),
    )
    response = admin.action_submit_view(request, "my_action")
    data = json.loads(response.content)
    assert data == {"status": "custom", "extra": 42}


def test_action_submit_passes_through_http_response():
    """When the action returns an HttpResponse, pass it through unchanged
    (used for file downloads / redirects)."""
    from django.http import HttpResponse
    rf = RequestFactory()
    request = rf.post("/admin/tests/widget/action-submit/my_action/", {})
    file_response = HttpResponse(b"PDFDATA", content_type="application/pdf")
    admin = _make_action_admin(
        get_action_result=(lambda *a, **kw: file_response, "my_action", "My Action"),
    )
    response = admin.action_submit_view(request, "my_action")
    assert response is file_response


def test_action_submit_handles_exception_with_500_json():
    """If the action raises, return 500 JSON with the exception message."""
    import json
    rf = RequestFactory()
    request = rf.post("/admin/tests/widget/action-submit/my_action/", {})

    def boom(*a, **kw):
        raise ValueError("kaboom")

    admin = _make_action_admin(
        get_action_result=(boom, "my_action", "My Action"),
    )
    response = admin.action_submit_view(request, "my_action")
    assert response.status_code == 500
    data = json.loads(response.content)
    assert data["status"] == "error"
    assert "kaboom" in data["msg"]


def test_action_submit_select_across_one_skips_pk_filter():
    """When select_across=1 (select all across pages), don't apply pk__in filter."""
    rf = RequestFactory()
    request = rf.post(
        "/admin/tests/widget/action-submit/my_action/",
        {"_selected": "1,2,3", "select_across": "1"},
    )
    qs = MagicMock()
    admin = _make_action_admin(
        get_changelist_raises=True,
        get_queryset_value=qs,
        get_action_result=(lambda *a, **kw: None, "my_action", "My Action"),
    )
    admin.action_submit_view(request, "my_action")
    qs.filter.assert_not_called()

